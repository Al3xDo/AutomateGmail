import openpyxl,datetime,smtplib,os,pprint
guestinfo={}
def timetransfer(time):
    return datetime.datetime.strptime(time+' 1','%b %Y %d')
# xử lý file excel
print('THE SOFT WILL USE THE FILE IN THE DIRECTORY '+ os.getcwd())
print('ONLY USE WITH GMAIL')
wb=openpyxl.load_workbook('duesRecords.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
lastr=sheet.max_row
lastc=sheet.max_column
#thiết lập ngày tháng
today=datetime.date.today()
now=datetime.datetime(int(today.year),int(today.month),1)
# tạo danh sách người nợ
for rowN in range(2,lastr+1):
    latestmon=sheet.cell(row=rowN,column=lastc).value
    guestname=sheet.cell(row=rowN,column=1).value
    guestemail=sheet.cell(row=rowN,column=2).value
    guestinfo.setdefault(guestname,{})
    guestinfo[guestname].setdefault(guestemail,{})
    dem='1'
    for col in range(3,lastc+1):
        if now>=timetransfer(sheet.cell(row=1,column=col).value) and sheet.cell(row=rowN,column=col).value!='paid':
            guestinfo[guestname][guestemail][dem]=sheet.cell(row=1,column=col).value
            dem=int(dem)+1
    if guestinfo[guestname][guestemail]=={}:
        del guestinfo[guestname]
# xử lý smtp
smtpObj=smtplib.SMTP('smtp.gmail.com',587)
if type(smtpObj)!=None:
    print('accesed sucessfully! ')
    smtpObj.ehlo()
    smtpObj.starttls()
    print('Your name account:')
    name=input()
    print('password:')
    mypass=input()
    # xử lý password
    try:
        smtpObj.login(name,mypass)
    except smtplib.SMTPAuthenticationError:
        print("Wrong account ID or you haven't allow our account use another program")
    print('Sending email.....')
    for guestname,guestemail in guestinfo.items():
        body='Subject: \n Dues unpaid reminding\nDear %s\nThis is an email reminding you to pay the credit after 15 days, you will need to pay extra fee for each dues unpaid, you need to pay the fee) \nI would be very grateful if you can pay it as soon as possible\nHere is your information\n%s\n\t\t\t\t\t\tSincerery\n\t\t\t\t\t\t  Jack'%(guestname,guestinfo[guestname])
        status=smtpObj.sendmail(name,guestemail,body)
        if status!={}:
            print('There was a problem sending email to %s: %s'%(guestemail,status))
    smtpObj.quit()
    print('Done')
